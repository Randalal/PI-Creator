import streamlit as st
import pandas as pd
from rapidfuzz import fuzz, process
from datetime import date   # 👈 就在这里加


st.set_page_config(page_title="产品检索工具 (Step 1)", layout="wide")
st.title("🔍 产品检索工具 (Step 1)")

# ---- 会话态：已选清单 ----
if "cart" not in st.session_state:
    st.session_state.cart = []  # list of dicts: {EAN, DESCRIPTION, QTY, RATE, AMOUNT}

def add_to_cart(row: pd.Series, qty: float):
    """将选中行加入清单；如已存在则合并数量并重算金额。"""
    ean = str(row["EAN"])
    desc = str(row["DESCRIPTION"])
    rate = float(row["RATE"])
    qty = float(qty)
    if qty <= 0:
        st.warning("数量必须大于 0")
        return

    # 若已存在同 EAN 的条目，则合并数量
    found = False
    for item in st.session_state.cart:
        if str(item["EAN"]) == ean:
            item["QTY"] += qty
            item["AMOUNT"] = round(item["QTY"] * item["RATE"], 2)
            found = True
            break
    if not found:
        st.session_state.cart.append({
            "EAN": ean,
            "DESCRIPTION": desc,
            "QTY": qty,
            "RATE": rate,
            "AMOUNT": round(qty * rate, 2),
        })
    st.success("已加入清单")

# ---- 上传价格表 ----
uploaded_file = st.file_uploader("上传价格表 (Excel/CSV)", type=["xlsx", "xls", "csv"])
df = None
required_cols = ["EAN", "DESCRIPTION", "RATE"]

if uploaded_file:
    try:
        if uploaded_file.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"读取文件失败：{e}")

# ---- 校验字段 ----
if df is not None:
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        st.error(f"缺少必要字段：{missing}。请保证表头包含 {required_cols}")
        df = None

# ---- 检索区 ----
if df is not None:
    st.divider()
    keyword = st.text_input("输入关键词进行检索（支持型号、描述、或 EAN 片段）")

    results = pd.DataFrame(columns=["EAN", "DESCRIPTION", "RATE"])
    if keyword:
        # 1) 在 DESCRIPTION 做模糊匹配（Top-N）
        desc_series = df["DESCRIPTION"].astype(str)
        fuzzy_matches = process.extract(
            keyword, desc_series, limit=50, scorer=fuzz.partial_ratio
        )
        idx_from_desc = [desc_series[desc_series == m[0]].index[0] for m in fuzzy_matches]

        # 2) 在 EAN 做包含匹配（便于直接输条码）
        ean_mask = df["EAN"].astype(str).str.contains(keyword, case=False, na=False)
        idx_from_ean = df[ean_mask].index.tolist()

        # 3) 合并去重
        all_idx = list(dict.fromkeys(idx_from_desc + idx_from_ean))
        results = df.loc[all_idx, ["EAN", "DESCRIPTION", "RATE"]].copy()

        # 选择 + 数量 + 加入清单（表格方式：data_editor + 复选框）
        if not results.empty:
            # 表格视图：加一列 SELECT，其他列只读
            results_view = results.reset_index(drop=True).copy()
            results_view.insert(0, "SELECT", False)

            picked = st.data_editor(
                results_view,
                hide_index=True,
                use_container_width=True,
                height=420,
                disabled=["EAN", "DESCRIPTION", "RATE"],  # 防止误改数据
                column_config={
                    "SELECT": st.column_config.CheckboxColumn("选择", help="勾选要加入的一行", default=False),
                    "EAN": st.column_config.TextColumn("EAN", width="medium"),
                    "DESCRIPTION": st.column_config.TextColumn("DESCRIPTION", width="large"),
                    "RATE": st.column_config.NumberColumn("RATE", format="%.2f"),
                },
                key="pick_editor",
            )

            # 读取被勾选的行
            checked = picked[picked["SELECT"]].drop(columns=["SELECT"])
            if len(checked) == 0:
                st.info("请在上面的表格勾选一行。")
            elif len(checked) > 1:
                st.warning("一次只能选择一行，请只勾选一条。")
            else:
                # 唯一选中行
                sel_row = checked.iloc[0]
                st.markdown(f'**已选择：** {sel_row["DESCRIPTION"]} | EAN:{sel_row["EAN"]} | RATE:{sel_row["RATE"]}')
                qty = st.number_input("输入数量（QTY）", min_value=0.0, step=1.0, value=1.0, key="qty_input")

                col_a, col_b = st.columns([1, 4])
                with col_a:
                    if st.button("➕ 加入清单", use_container_width=True):
                        add_to_cart(sel_row, qty)
                with col_b:
                    st.caption("提示：可继续在上方输入下一个关键词，重复添加。")

    # ---- 已选清单展示（信息行） ----
    st.divider()
    st.subheader("🧾 已选清单（信息行）")
    if st.session_state.cart:
        cart_df = pd.DataFrame(st.session_state.cart, columns=["EAN", "DESCRIPTION", "QTY", "RATE", "AMOUNT"])
        st.dataframe(cart_df, use_container_width=True)

        # 简单的删除功能（可选）
        with st.expander("删除某条（可选）"):
            ean_to_remove = st.text_input("输入要删除的 EAN")
            if st.button("删除"):
                before = len(st.session_state.cart)
                st.session_state.cart = [x for x in st.session_state.cart if str(x["EAN"]) != str(ean_to_remove)]
                after = len(st.session_state.cart)
                if after < before:
                    st.success("已删除")
                else:
                    st.warning("未找到该 EAN")
    else:
        st.info("清单为空。请先在上方检索并加入。")
else:
    st.info("请先上传价格表（需包含列：EAN, DESCRIPTION, RATE）。")





st.title("📄 PI 生成工具 (Step 2)")

# ---- 上传 PI 模版 ----
pi_template_file = st.file_uploader(
    "上传 PI 模版 (Excel, 仅支持 .xlsx)",
    type=["xlsx"],
    key="pi_template_uploader"
)

if pi_template_file:
    st.session_state["pi_template"] = pi_template_file
    st.success(f"已加载模版文件：{pi_template_file.name}")
else:
    st.info("请上传 PI 模版 (.xlsx)，作为生成 PI 的基础。")

# ================== Step 3：写入 PI 模版并导出（固定坐标） ==================
import io
from openpyxl import load_workbook

st.divider()
st.header("📤 Step 3：写入 PI 模版并导出")

# 1) 确保已有模版（来自 Step 2 上传）
if "pi_template" not in st.session_state:
    st.info("请先到 Step 2 上传 PI 模版（.xlsx）")
    st.stop()

# 2) 确保有已选清单
if not st.session_state.get("cart"):
    st.info("清单为空。请先在 Step 1 检索并加入产品。")
    st.stop()

# 3) 组装清单 DataFrame 与合计
cart_df = pd.DataFrame(
    st.session_state.cart,
    columns=["EAN", "DESCRIPTION", "QTY", "RATE", "AMOUNT"]
).copy()

# ==== 用户输入文件名（可选），默认值为 PIM{yymmdd}KSA001 ====
today = date.today()
yymmdd = today.strftime("%y%m%d")
region_suffix = "KSA001"  # 你们固定的尾缀，可改为你自己的
default_pi_no = f"PIM{yymmdd}{region_suffix}"

filename_input = st.text_input("请输入导出的 PI 文件名（不含扩展名）", value=default_pi_no)
pi_no = filename_input.strip() or default_pi_no

# 数值化，避免字符串导致写入失败
cart_df["QTY"] = pd.to_numeric(cart_df["QTY"], errors="coerce").fillna(0.0)
cart_df["RATE"] = pd.to_numeric(cart_df["RATE"], errors="coerce").fillna(0.0)
cart_df["AMOUNT"] = (cart_df["QTY"] * cart_df["RATE"]).round(2)

qty_total   = float(cart_df["QTY"].sum())
subtotal    = float(cart_df["AMOUNT"].sum())
grand_total = subtotal  # 如需含税/加费，这里调整公式

# 4) 导出按钮
if st.button("生成并下载 PI（Excel）", type="primary", key="export_pi_btn"):
    # 从 session_state 的上传文件读取为 BytesIO，再交给 openpyxl
    tpl_buf = io.BytesIO(st.session_state["pi_template"].getvalue())
    wb = load_workbook(tpl_buf)
    ws = wb["Invoice"] if "Invoice" in wb.sheetnames else wb.active

    # --- 根据当天生成日期与 PI 号，并写入抬头 ---
    date_str = today.strftime("%Y-%m-%d")  # 写入 E3 的日期
    ws["E3"] = f" {date_str}"  # 修改 E3
    ws["E4"] = f"{pi_no}"  # 结果: INVOICE #PIM250827KSA001

    # 5) 写入明细（固定坐标：从第14行起，A..E）
    start_row = 14
    for i, row in cart_df.iterrows():
        r = start_row + i
        ws[f"A{r}"] = str(row["EAN"])
        ws[f"B{r}"] = str(row["DESCRIPTION"])
        ws[f"C{r}"] = float(row["QTY"])
        ws[f"D{r}"] = float(row["RATE"])
        ws[f"E{r}"] = float(row["AMOUNT"])

    # 6) 写入合计（固定坐标）
    ws["C31"] = qty_total
    ws["E31"] = subtotal
    ws["E34"] = grand_total

    # 7) 保存到内存并提供下载
    out = io.BytesIO()
    wb.save(out)
    st.download_button(
        label="⬇️ 下载 Excel",
        data=out.getvalue(),
        file_name=f"{pi_no}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_pi_btn"
    )
    st.success("已根据模版生成 Excel。")

